from dataclasses import dataclass
import logging
import requests
from bs4 import BeautifulSoup
import xlsxwriter
import zipfile
import os
import openpyxl
from openpyxl import load_workbook,Workbook
import csv
import dask.dataframe as dd

if not os.path.exists("log"):
    os.makedirs("log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("log/debug.log",mode='w+'),
        logging.StreamHandler()
    ]
)
import urllib3
#This is to disable the request error due to certifigate verification 
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

@dataclass
class BaseModule:
    """
    This BaseModule Class implements the entire logic required to scrape the specified url and perform the data transformation 
    ** Use of dataclass ensures we don't need to add the self.__init__() method explicitly.
    """
    excel_data_list = []

    def init_extract_transform(self, url, total_page_length,per_page_data):
        """
        This method is the entrypoint method to start the entire scrape and transformation process.
        """
        logging.info(f"Starting the process to scrape date from the following url {url} and cretion of the Excel-sheet")
        for i in range(1, total_page_length+1):
            page = requests.get(
                f"{url}/?page_num={i}&per_page={per_page_data}", verify=False)
            soup = BeautifulSoup(page.text, 'html.parser')
            table = soup.find('table', {"class": "table"})
            file = open(f"data/html_data/{i}.html", "w")
            file.write(str(table))
            file.close()
            self._scrape_data(table)
        self._create_zip_from_html(total_page_length)
        self._raw_to_xl_workbook()
        self._create_dataset()
        self._add_transformed_data_to_xl()

    def _scrape_data(self, table):
        """
        This method scrapes the page for the specific HTML tag and retieves the data and stores in a python list.
        """
        try:
            for j in table.find_all('tr')[1:]:
                row_data = j.find_all('td')
                row = [i.text for i in row_data]
                self.excel_data_list.append(row)
        except Exception as e:
            logging.exception(e)

    def _raw_to_xl_workbook(self):
        """
        This method is to create a Excel Workbook with the scraped data
        """
        logging.info(f"Creating the NHL Stats 1990-2011 worksheet in scraped_excel_data.xlsx")
        # Create an new Excel file and add a worksheet.
        workbook = xlsxwriter.Workbook(
            'data/excel_data/scraped_excel_data.xlsx')
        worksheet = workbook.add_worksheet()

        worksheet.add_table("A1:I2500",
                            {'data': self.excel_data_list,
                             'columns': [
                                 {'header': ' Team Name '},
                                 {'header': ' Year '},
                                 {'header': ' Wins '},
                                 {'header': ' Losses '},
                                 {'header': ' OT Losses '},
                                 {'header': ' Win %'},
                                 {'header': ' Goals For (GF) '},
                                 {'header': ' Goals Against (GA) '},
                                 {'header': ' + / -'}]
                             })

        worksheet.set_column(0, 9, 30)
        workbook.close()

    def _create_zip_from_html(self, total_page_length):
        """
        This method is to create the Zip file with the scraped HTML data 
        """
        logging.info("Creating the ZIP file from the scraped HTML files")
        if not os.path.exists('data/zip_data'):
            logging.info(
                f"creating the following directory- data/zip_data to store the zipped html-file")
            os.makedirs('data/zip_data')
        else:
            os.remove("data/zip_data/html_zip.zip")
            logging.info("directory data/zip_data already exists")

        for i in range(1, total_page_length+1):
            with zipfile.ZipFile("data/zip_data/html_zip.zip", mode="a") as archive:
                archive.write(f"data/html_data/{i}.html")

    def _create_dataset(self):
        """
        This method creates a dataframe in csv format to be consumed in DASK for the transformation step
        """
        logging.info("Creating a dataset from the scraped data to be consumed in DASK fro the transformation process")
        # Define variable to load the dataframe
        wb = openpyxl.load_workbook('data/excel_data/scraped_excel_data.xlsx')
        sh = wb.active
        if not os.path.exists('data/dataset'):
            logging.info(
                f"creating the following directory- data/zip_data to store the zipped html-file")
            os.makedirs('data/dataset')
        else:
            logging.info("directory data/dataset already exists")
        with open('data/dataset/dataset.csv', 'w', newline="") as f:
            c = csv.writer(f)
            for r in sh.rows:
                c.writerow([cell.value for cell in r])

    def _add_transformed_data_to_xl(self):
        """
        This method reads the dataset created into a DASK dataframe and then applies a groupby to identify the Winners and Losers by Year
        """
        logging.info("Starting the transformation process to consume the dataset in DASK and do a groupby to get the winners and losers by year")
        try:
            df = dd.read_csv('data/dataset/dataset.csv')
            df.columns = df.columns.str.strip()

            df_max = df.groupby(["Year"])["Wins"].agg('max').reset_index().compute()
            df_min = df.groupby(["Year"])["Wins"].agg('min').reset_index().compute()

            list_data_max=[]
            list_data_min=[]
            for i in range(df_max['Year'].count()):
                val = df[['Year','Team Name','Wins']].loc[(df["Year"] == df_max['Year'][i]) & (df["Wins"]==df_max['Wins'][i])].reset_index(drop=True).values.compute()
                list_data_max.append(val[0])
                val= df[['Year','Team Name','Wins']].loc[(df["Year"] == df_min['Year'][i]) & (df["Wins"]==df_min['Wins'][i])].reset_index(drop=True).values.compute()
                list_data_min.append(val[0])

            
            ## adding the result to the existing excel sheet

            book = load_workbook('data/excel_data/scraped_excel_data.xlsx')
            book.create_sheet('Winner and Loser per Year',1)
            sheet = book["Winner and Loser per Year"]

            sheet['A1'] = "Year"
            sheet['B1'] = "Winner"
            sheet['C1'] = "Winner Num. of Wins"
            sheet['D1'] = "Loser"
            sheet['E1'] = "Loser Num. of Wins"
            
            column_list = ["B","C","D","E"]
            for i in column_list:
                sheet.column_dimensions[i].width = 30

            for i in range(len(list_data_max)):
                sheet[f"A{i+2}"]= list_data_max[i][0]
                sheet[f"B{i+2}"]= list_data_max[i][1]
                sheet[f"C{i+2}"]= list_data_max[i][2]
                sheet[f"D{i+2}"]= list_data_min[i][1]
                sheet[f"E{i+2}"]= list_data_min[i][2]
            
            sheet = book["Sheet1"]
            sheet.title = "NHL Stats 1990-2011"
            book.save('data/excel_data/scraped_excel_data.xlsx')

            book.close()

        except Exception as e:
            logging.exception(e)


